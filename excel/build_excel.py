"""Build excel/FinSight_Analytics.xlsx — a small analyst-skill demonstration.

Sheets:
  Raw_Data          -- a ~500-row enriched sample of transactions
  Pivot_Analysis    -- category x month revenue cross-tab (SUMIFS) + bar chart
  Campaign_Analysis -- campaign performance with conditional formatting on ROI
  Customer_Analysis -- per-segment rollups via SUMIFS/COUNTIFS + XLOOKUP demo
  Summary           -- KPI cards referencing the other sheets

Demonstrates: XLOOKUP, SUMIFS, COUNTIFS, formula-driven cross-tab (pivot-style),
a chart, and conditional formatting. A native Excel PivotTable can be added on
Raw_Data in one step (Insert > PivotTable); this workbook shows the equivalent
with formulas so the file is fully reproducible from code.

Run:  python excel/build_excel.py
"""
from __future__ import annotations

from pathlib import Path
import pandas as pd

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
PBI = ROOT / "outputs" / "powerbi"
OUT = ROOT / "excel" / "FinSight_Analytics.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _write_df(ws, df, start_row=1, start_col=1):
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=str(col))
    for i, (_, r) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=start_col):
            v = r[col]
            ws.cell(row=i, column=j, value=(None if pd.isna(v) else v))
    _style_header(ws, start_row, len(df.columns))
    for j, col in enumerate(df.columns, start=start_col):
        width = min(max(len(str(col)) + 2, 12), 28)
        ws.column_dimensions[get_column_letter(j)].width = width


def build():
    fact = pd.read_csv(PBI / "fact_transactions.csv")
    camp = pd.read_csv(PBI / "campaign_performance.csv")
    rfm = pd.read_csv(PBI / "rfm_summary.csv")
    cats = sorted(fact["category"].dropna().unique())
    months = sorted(fact["month"].dropna().unique())

    # sample raw data (seeded for reproducibility)
    sample = fact.sample(min(500, len(fact)), random_state=42).sort_values("transaction_date")
    sample_cols = ["transaction_id", "customer_id", "transaction_date", "month",
                   "category", "merchant_name", "transaction_amount",
                   "is_campaign_transaction", "income_band", "rfm_segment", "city"]
    sample = sample[sample_cols].reset_index(drop=True)

    wb = Workbook()

    # ---------- Raw_Data ----------
    ws = wb.active
    ws.title = "Raw_Data"
    ws["A1"] = "FinSight — Raw Transaction Sample (synthetic)"
    ws["A1"].font = TITLE_FONT
    _write_df(ws, sample, start_row=3)
    n = len(sample)
    data_first, data_last = 4, 3 + n
    amt_col = sample_cols.index("transaction_amount") + 1  # 1-based
    amt_letter = get_column_letter(amt_col)
    cat_letter = get_column_letter(sample_cols.index("category") + 1)
    camp_letter = get_column_letter(sample_cols.index("is_campaign_transaction") + 1)
    seg_letter = get_column_letter(sample_cols.index("rfm_segment") + 1)
    # colour scale on the amount column
    ws.conditional_formatting.add(
        f"{amt_letter}{data_first}:{amt_letter}{data_last}",
        ColorScaleRule(start_type="min", start_color="FFF8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
                       end_type="max", end_color="FF63BE7B"))
    ws.freeze_panes = "A4"

    # named ranges (as absolute refs used in formulas below)
    rng_amt = f"Raw_Data!${amt_letter}${data_first}:${amt_letter}${data_last}"
    rng_cat = f"Raw_Data!${cat_letter}${data_first}:${cat_letter}${data_last}"
    rng_camp = f"Raw_Data!${camp_letter}${data_first}:${camp_letter}${data_last}"
    rng_seg = f"Raw_Data!${seg_letter}${data_first}:${seg_letter}${data_last}"
    month_col_letter = get_column_letter(sample_cols.index("month") + 1)
    rng_month = f"Raw_Data!${month_col_letter}${data_first}:${month_col_letter}${data_last}"

    # ---------- Pivot_Analysis: category x month revenue via SUMIFS ----------
    ws = wb.create_sheet("Pivot_Analysis")
    ws["A1"] = "Revenue by Category x Month (SUMIFS over Raw_Data — pivot-style)"
    ws["A1"].font = TITLE_FONT
    ws.cell(row=3, column=1, value="Category")
    for j, mth in enumerate(months, start=2):
        ws.cell(row=3, column=j, value=mth)
    ws.cell(row=3, column=len(months) + 2, value="Total")
    _style_header(ws, 3, len(months) + 2)
    for i, cat in enumerate(cats, start=4):
        ws.cell(row=i, column=1, value=cat)
        for j, mth in enumerate(months, start=2):
            col_letter = get_column_letter(j)
            f = f'=SUMIFS({rng_amt},{rng_cat},$A{i},{rng_month},{col_letter}$3)'
            ws.cell(row=i, column=j, value=f)
        tot_col = len(months) + 2
        first = get_column_letter(2)
        last = get_column_letter(len(months) + 1)
        ws.cell(row=i, column=tot_col, value=f"=SUM({first}{i}:{last}{i})")
    ws.column_dimensions["A"].width = 16
    # bar chart of category totals (pivot-chart substitute)
    chart = BarChart()
    chart.title = "Sample Revenue by Category"
    chart.type = "col"
    chart.height, chart.width = 9, 18
    tot_col_letter = get_column_letter(len(months) + 2)
    data_ref = Reference(ws, min_col=len(months) + 2, min_row=3, max_row=3 + len(cats))
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + len(cats))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{6 + len(cats)}")

    # ---------- Campaign_Analysis ----------
    ws = wb.create_sheet("Campaign_Analysis")
    ws["A1"] = "Campaign Performance"
    ws["A1"].font = TITLE_FONT
    camp_cols = ["campaign_id", "campaign_name", "campaign_category", "channel",
                 "engaged", "redeemed", "redemption_rate", "campaign_revenue",
                 "discount_cost", "roi"]
    cdf = camp[camp_cols].copy()
    _write_df(ws, cdf, start_row=3)
    c_first, c_last = 4, 3 + len(cdf)
    roi_letter = get_column_letter(camp_cols.index("roi") + 1)
    red_letter = get_column_letter(camp_cols.index("redemption_rate") + 1)
    # highlight negative ROI in red, strong redemption in green
    ws.conditional_formatting.add(
        f"{roi_letter}{c_first}:{roi_letter}{c_last}",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006")))
    ws.conditional_formatting.add(
        f"{red_letter}{c_first}:{red_letter}{c_last}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))
    ws.freeze_panes = "A4"

    # ---------- Customer_Analysis ----------
    ws = wb.create_sheet("Customer_Analysis")
    ws["A1"] = "Customer Segment Analysis (SUMIFS / COUNTIFS)"
    ws["A1"].font = TITLE_FONT
    ws.cell(row=3, column=1, value="RFM Segment")
    ws.cell(row=3, column=2, value="Transactions (COUNTIFS)")
    ws.cell(row=3, column=3, value="Revenue (SUMIFS)")
    ws.cell(row=3, column=4, value="Campaign Txns (COUNTIFS)")
    _style_header(ws, 3, 4)
    segments = sorted(fact["rfm_segment"].dropna().unique())
    for i, seg in enumerate(segments, start=4):
        ws.cell(row=i, column=1, value=seg)
        ws.cell(row=i, column=2, value=f'=COUNTIFS({rng_seg},$A{i})')
        ws.cell(row=i, column=3, value=f'=SUMIFS({rng_amt},{rng_seg},$A{i})')
        ws.cell(row=i, column=4, value=f'=COUNTIFS({rng_seg},$A{i},{rng_camp},1)')
    ws.column_dimensions["A"].width = 16
    for col in "BCD":
        ws.column_dimensions[col].width = 22

    # XLOOKUP demonstration: look up a customer's income band from dim_customers
    dim = pd.read_csv(PBI / "dim_customers.csv")[["customer_id", "income_band", "rfm_segment"]]
    dws = wb.create_sheet("dim_customers")
    _write_df(dws, dim, start_row=1)
    d_last = 1 + len(dim)
    ws.cell(row=4 + len(segments) + 1, column=1, value="XLOOKUP demo — income band for a customer id:")
    ws.cell(row=4 + len(segments) + 2, column=1, value="Customer ID")
    ws.cell(row=4 + len(segments) + 2, column=2, value=str(dim["customer_id"].iloc[0]))
    ws.cell(row=4 + len(segments) + 3, column=1, value="Income band")
    ws.cell(row=4 + len(segments) + 3, column=2,
            value=(f'=XLOOKUP(B{4 + len(segments) + 2},'
                   f'dim_customers!$A$2:$A${d_last},dim_customers!$B$2:$B${d_last},"Not found")'))

    # ---------- Summary ----------
    ws = wb.create_sheet("Summary")
    ws["A1"] = "FinSight — Summary (synthetic data)"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    kpis = [
        ("Total sample transactions", f'=COUNTA({rng_cat})'),
        ("Total sample revenue", f'=SUM({rng_amt})'),
        ("Campaign transactions (sample)", f'=COUNTIFS({rng_camp},1)'),
        ("Avg transaction value (sample)", f'=AVERAGE({rng_amt})'),
    ]
    for i, (label, formula) in enumerate(kpis, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=formula)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws["A9"] = ("Note: figures on this sheet summarise the 500-row Raw_Data sample, "
                "not the full dataset. All data is synthetic.")
    ws["A9"].font = Font(italic=True, color="808080")

    # order sheets sensibly
    wb.move_sheet("dim_customers", offset=len(wb.sheetnames))
    wb.save(OUT)
    print(f"Workbook written to {OUT}")
    print(f"  Raw_Data sample rows: {n}")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    build()
